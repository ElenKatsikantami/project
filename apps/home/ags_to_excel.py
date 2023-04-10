import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
from python_ags4 import AGS4
import pandas as pd
import numpy as np
import os
import re
import pyproj
import io
from sqlalchemy import create_engine
from django.conf import settings
import datetime
import openpyxl
from openpyxl.styles import PatternFill ,Font , Alignment

def get_rock_level(df_GEOL,id_name,areas):
    rock_level = {}
    for id_ in areas:
        value = list(df_GEOL[(df_GEOL[id_name] == id_) & 
                            ((df_GEOL["GEOL_DESC"].str.find("SANDSTONE")>0)| 
                            (df_GEOL["GEOL_DESC"].str.find("CALCARENITE")>0)
#                               |(df_GEOL["GEOL_DESC"].str.find("any future word")>0)
                            )]["GEOL_TOP"].head(1))

        if len(value) ==0:
            rock_level[id_] = "NA"
        else:
            rock_level[id_] = value[0]  
    return [rock_level[id_] for id_ in areas]

def get_summery(ags_edited_path,id_name):
    tables, _ = AGS4.AGS4_to_dataframe(ags_edited_path)
    tables_delete=["FILE","MONG","PTIM"]
    # combine some sheets in one (which have criteria of the boreholes)
    tables_sum = []
    for table in tables.keys():
        if id_name in tables[table].columns:
            if len(tables[table][id_name].unique())==len(tables[table][id_name]) and table not in tables_delete:
                tables_sum.append(table)
                
    tables_Rest = [i for i in tables.keys() if i not in (tables_delete+tables_sum)]
    # take the first sheet as a core of the dataframe 
    df = tables[tables_sum[0]]
    df = df.replace(["X","DATA"],np.nan)
    tables_sum.pop(0)

    # merging the rest of the tables
    for table in tables_sum:
        df = pd.merge(df,tables[table], on=[id_name] ,how="outer",suffixes=(f'_{table}', f'_{table}2'))
        df = df.replace(["X","DATA"],np.nan)

    # re-arranging the collumns order 
    columnsNames = sorted(df.columns)
    columnsNames.remove(id_name)
    columnsNames.insert(0,id_name)
    df=df.loc[:, columnsNames]

    # drop the first row to have unique typy of data in each column
    if id_name != 'LOCA_ID':
        df = df.drop(0)
    else:
        df = df.drop([0,1])
        
    # delete repeated columns and remove empty ones
    df = df.T.drop_duplicates().T
    df = df.replace([""],np.nan)
    df = df.dropna(axis=1, thresh=2)

    # turn data to numeric
    df = df.replace(["-"],0)
    for column in df.columns:
        df[column] = pd.to_numeric(df[column],errors='ignore')
    try:
    # turn AGS to WGS48
        location = " ".join(tables["PROJ"]["PROJ_LOC"])
        if "Saudi Arabia" in location:
            EPSG = 32636
        if "Abu Dhabi" in location:
            EPSG = 32640
        if "Dubai" in location:
            EPSG = 3997
        coordinates = ["LOCA_LOCX","LOCA_LOCY","LOCA_NATE","LOCA_NATN","HOLE_NATE","HOLE_NATN"]
        coordinates_columns = [i for i in df.columns if i in coordinates]
        leigon_wgs = pyproj.Transformer.from_crs(EPSG,4326)
        df["latitude"],df["longitude"] = leigon_wgs.transform(df[coordinates_columns[0]],df[coordinates_columns[1]])
    except:
        coordinates_columns=[]
    
    sheets= {"BH SUM":df }
    for table in tables_Rest:
        exec(f"""df_{table} = tables[table]
if id_name != 'LOCA_ID':
    df_{table} = df_{table}.drop(0)
else:
    df_{table} = df_{table}.drop([0,1])
df_{table} = df_{table}.replace(["-"],0)
for column in df_{table}.columns:
    df_{table}[column] = pd.to_numeric(df_{table}[column],errors='ignore')
sheets[table] = df_{table}""")
    return sheets,coordinates_columns

def get_main_df(sheets,id_name,coordinates_columns) :
    df = sheets["BH SUM"]
    elevations = re.findall("\w+_GL|\w+_LOCZ"," ".join(df.columns))
    groundWater = ["WSTG_DPTH","MOND_RDNG","WSTK_DEP"]
    groundWater_column = [i for i in df.columns if i in groundWater]
    
    starting_columns = [id_name]+coordinates_columns+["latitude","longitude"]+elevations+groundWater_column+["HDPH_EXC"]
    if len(df[id_name]) == len(df[groundWater_column[0]].dropna()):
        df = df[starting_columns].copy().dropna(axis=1)
    else:
        df = df[starting_columns].copy().dropna()
        
    df["Rock Level"] = get_rock_level(sheets['GEOL'],id_name,df[id_name].unique())    
        
    elevation = re.findall("\w+_GL|\w+_LOCZ"," ".join(df.columns))[0]

    df = df.set_index(id_name)
    return df , elevation

def get_df(sheet,Columns,sheets,id_name,df,elevation,ags_file):
        """
        this function for making the dataframe and add elevation for it
        returns dataframe
        """
        if sheet not in sheets:
            return
        if sheet == "CORE":
            columns_0 = Columns+["CORE_TOP","CORE_BASE",id_name]
        elif sheet == "ISPT":
            columns_0 = Columns+["ISPT_TOP",id_name]
        else:
            columns_0 = Columns+["SAMP_TOP",id_name]     
        
        df_0 = sheets[sheet][columns_0].set_index(id_name)
        
        for column in df_0.columns:
                df_0[column] = pd.to_numeric(df_0[column],errors='ignore')
        
        df_0["Loca_gl"] = df[elevation]
        df_0["machine"] = df["HDPH_EXC"]
        df_0 = df_0.reset_index()
        
        if sheet == "CORE":
            df_0["depth"] = df_0[["CORE_TOP","CORE_BASE"]].mean(axis=1)
        elif sheet == "ISPT":
            df_0["depth"] = df_0["ISPT_TOP"]
            if df_0["ISPT_NVAL"].dtype == "O":
                df_0["ISPT_NVAL"]= df_0["ISPT_NVAL"].apply(lambda x: x.split("/")[0] if not x.isalnum() else x)
                df_0["ISPT_NVAL"]=pd.to_numeric(df_0["ISPT_NVAL"])
        else:
            df_0["depth"] = df_0["SAMP_TOP"]
            
        df_0["elevation"] = df_0.Loca_gl - df_0.depth
        df_0["borhole_id"] = df_0[id_name]
        df_0["ags_file_id"] = ags_file.id
        df_0 = df_0[["borhole_id","machine","depth","elevation"]+Columns+["ags_file_id"]].fillna(0)
        
        return df_0   

def ags_to_excel(ags_file, summery,info,format,record=False):
    path = os.path.join(os.path.dirname(ags_file.ags_file.path),"edited_ags")
    isExist = os.path.exists(path)
    if not isExist:
        os.makedirs(path)

    #editing the ags file
    id_name = 'LOCA_ID'
    with open(ags_file.ags_file.path,"r") as file:
        lines = file.readlines()
        if lines[0].find("**")+1:
            id_name ="HOLE_ID"
            for i in range(len(lines)):
                if len(lines[i]) > 1:
                    if lines[i].find("**")+1:
                        lines[i]='"GROUP",'+f'"{lines[i][3:-2]}"'+'\n'
                    if lines[i].find("*")+1:
                        lines[i]=lines[i].replace("*","")
                    if lines[i-1].split(",")[0] == '"GROUP"':
                        lines[i]='"HEADING",'+lines[i]
                    if lines[i-1].split(",")[0] == '"HEADING"' or lines[i-1].split(",")[0] == '"DATA"':
                        lines[i]='"DATA",'+lines[i]
        else:
            for i in range(len(lines)):
                if len(lines[i]) > 1:
                    if lines[i][-2] != '"':
                        lines[i]=lines[i].strip() + " "

    # storing the edited ones
    name = os.path.split(ags_file.ags_file.path)[1]
    ags_edited_path = os.path.join(path,name)
    with open(ags_edited_path,"w+") as edited_file:
        edited_file.writelines(lines)
        
    isExist = os.path.exists(os.path.join("media","project","excel"))
    if not isExist:
        os.makedirs(os.path.join("media","project","excel"))
    
    sheets,coordinates_columns = get_summery(ags_edited_path,id_name)
    Project_ID = sheets["PROJ"]["PROJ_ID"].iloc[-1]
    Project_Name = sheets["PROJ"]["PROJ_NAME"].iloc[-1]
    Project_Location = sheets["PROJ"]["PROJ_LOC"].iloc[-1]
    Client = sheets["PROJ"]["PROJ_CLNT"].iloc[-1]
    try:
        Originator = sheets["PROJ"]["PROJ_CONT"].iloc[-1]
    except:
        Originator = sheets["PROJ"]["PROJ_ENG"].iloc[-1]
    
    df = sheets["BH SUM"]
    elevations = re.findall("\w+_GL|\w+_LOCZ"," ".join(df.columns))
    groundWater = ["WSTG_DPTH","MOND_RDNG","WSTK_DEP"]
    groundWater_column = [i for i in df.columns if i in groundWater]
    N_Borehole = df[id_name].count()
    Average_Ground_Level = np.round(max([np.average(df[elevation]) for elevation in elevations]),2)
    Average_Water_Level = np.round(np.average(df[groundWater_column]),2)
    if id_name != 'LOCA_ID':
        Max_Borehole_depth_cols = ["HOLE_FDEP","DOBS_BASE","HORN_BASE"]
    else:
        Max_Borehole_depth_cols = ["LOCA_FDEP","DOBS_BASE","HORN_BASE"]
    Max_Borehole_depth_col = [i for i in df.columns if i in Max_Borehole_depth_cols][0]
    Max_Borehole_depth= df[Max_Borehole_depth_col].max()
    total_drilled = df[Max_Borehole_depth_col].sum()
    
    if summery:
        if format == "Excel":
            excel_path_edited = os.path.join("media","project","excel",f"{name[:-4]}_edited.xlsx")
            file_path = excel_path_edited
            writer = pd.ExcelWriter(excel_path_edited)
            for sheet_name,dataframe in sheets.items():
                dataframe.to_excel(writer, sheet_name = sheet_name,index =False,encoding="utf-8")
            writer.close()
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.create_sheet(title="Statical Summery", index=1)
            row = 1
            fill_cell = PatternFill(patternType='solid', fgColor='FFFF00')
            sheet.cell(row = row, column = 1).value = "Field Test Data"
            sheet.cell(row = row, column = 2).value = "Value"
            sheet.cell(row = row, column = 3).value = "Units"
            sheet.cell(row = row, column = 1).font = Font(bold = True )
            for col in range(1, 10):
                    cell_header = sheet.cell(row, col)
                    cell_header.fill = fill_cell  
            row +=1

            Field_Test_Data = [("CORE_PREC" , sheets["CORE"]["CORE_PREC"].count()-1-int(id_name == 'LOCA_ID'),"Number")
            ,("CORE_SREC" , sheets["CORE"]["CORE_SREC"].count()-1-int(id_name == 'LOCA_ID'),"Number")
            ,("CORE_RQD" , sheets["CORE"]["CORE_RQD"].count()-1-int(id_name == 'LOCA_ID'),"Number")
            ,("N SPT Data" , sheets["ISPT"]["ISPT_NVAL"].count()-1-int(id_name == 'LOCA_ID'),"Number")]
            for i in range(len(Field_Test_Data)):
                sheet.cell(row = row, column = 1).value = Field_Test_Data[i][0]
                sheet.cell(row = row, column = 2).value = Field_Test_Data[i][1]
                sheet.cell(row = row, column = 3).value = Field_Test_Data[i][2]
                row +=1
            Laboratory = ['GRAG','LLPL','LNMC','LPDN','RDEN','RELD','TREG','RPLT','RUCS','SHBG',"CLSS"]
            test_names = {'GRAG':"Gradation Tests",'LLPL':"Aterrberg Tests",'CLSS':"Aterrberg Tests",
            'LNMC':"Water/Moisture Content",'LPDN':"Specific Gravity",
            'RDEN':"Bulk Density",'RELD':"Relative Density",
            'TREG':"Triaxial Tests (Effective Stress)",'RPLT':"Point Load Testing",
            'RUCS':"UCS Tests",'SHBG':"Shear Box Testing"}
            test_sheets = [i for i in sheets.keys() if i in Laboratory]
            Laboratory_Test_Data =[]
            
            if id_name == 'LOCA_ID':
                for i in test_sheets:
                    Laboratory_Test_Data.append([test_names[i], sheets[i]["SAMP_TOP"].count()-2,
                                                "Number",sheets[i][f"{i}_METH"].iloc[-1]])
            else:
                for i in test_sheets:
                    Laboratory_Test_Data.append([test_names[i], sheets[i]["SAMP_TOP"].count()-1,"Number",""])
                if "ROCK" in sheets.keys():
                    Laboratory_Test_Data.append(["Water/Moisture Content", sheets['ROCK']["ROCK_MC"].replace("",np.nan).count()-1,
                                                "Number",""])
                    Laboratory_Test_Data.append(["UCS Tests", sheets['ROCK']["ROCK_UCS"].replace("",np.nan).count()-1,
                                                "Number",""])
                    Laboratory_Test_Data.append(["Bulk Density", sheets['ROCK']["ROCK_BDEN"].replace("",np.nan).count()-1,
                                                "Number",""])
                    

            chemical_tests = {"CL" : "Chloride Content","CO2": "Carbonate Content" ,"PH":"PH Value", "WS" : "Sulphate Content"
                            ,"PHS":"PH Value",'CACO3': "Carbonate Content", "SO4" : "Sulphate Content"}
            Chemical_Test_Data = []
            if 'GCHM' in sheets.keys():
                for i in ["CL"  , "WS" , "CO2", "PH"]:
                    Chemical_Test_Data.append([chemical_tests[i] , sheets['GCHM'][sheets['GCHM']["GCHM_CODE"]==i]["GCHM_CODE"].count()
                                            ,"Number",sheets['GCHM']["GCHM_METH"].iloc[-1]])
            if 'CNMT' in sheets.keys():
                for i in ['CACO3', 'CL', 'PHS', 'SO4']:
                    Chemical_Test_Data.append([chemical_tests[i] , sheets['CNMT'][sheets['CNMT']["CNMT_TYPE"]==i]["CNMT_TYPE"].count()
                                            ,"Number",""])
            sheet.cell(row = row, column = 1).value = "Laboratory Test Data"
            sheet.cell(row = row, column = 1).font = Font(bold = True )
            sheet.cell(row = row, column = 2).value = "Value"
            sheet.cell(row = row, column = 2).font = Font(bold = True )
            sheet.cell(row = row, column = 3).value = "Units"
            sheet.cell(row = row, column = 3).font = Font(bold = True )
            sheet.cell(row = row, column = 4).value = "Test reference"
            sheet.cell(row = row, column = 4).font = Font(bold = True )

            for col in range(1, 10):
                    cell_header = sheet.cell(row, col)
                    cell_header.fill = fill_cell  
            row +=1
            for i in range(len(Laboratory_Test_Data)):
                sheet.cell(row = row, column = 1).value = Laboratory_Test_Data[i][0]
                sheet.cell(row = row, column = 2).value = Laboratory_Test_Data[i][1]
                sheet.cell(row = row, column = 3).value = Laboratory_Test_Data[i][2]
                sheet.cell(row = row, column = 4).value = Laboratory_Test_Data[i][3]
                row +=1

            sheet.cell(row = row, column = 1).value = "Chemical Test Data"
            sheet.cell(row = row, column = 1).font = Font(bold = True )
            sheet.cell(row = row, column = 2).value = "Value"
            sheet.cell(row = row, column = 2).font = Font(bold = True )
            sheet.cell(row = row, column = 3).value = "Units"
            sheet.cell(row = row, column = 3).font = Font(bold = True )
            sheet.cell(row = row, column = 4).value = "Test reference"
            sheet.cell(row = row, column = 4).font = Font(bold = True )

            for col in range(1, 10):
                    cell_header = sheet.cell(row, col)
                    cell_header.fill = fill_cell  
            row +=1
            for i in range(len(Chemical_Test_Data)):
                sheet.cell(row = row, column = 1).value = Chemical_Test_Data[i][0]
                sheet.cell(row = row, column = 2).value = Chemical_Test_Data[i][1]
                sheet.cell(row = row, column = 3).value = Chemical_Test_Data[i][2]
                sheet.cell(row = row, column = 4).value = Chemical_Test_Data[i][3]
                row +=1
            sheet.column_dimensions['A'].width = 37
            sheet.column_dimensions['B'].width = 40
            for row in range(1, 20):
                    cell_header = sheet.cell(row, 2)
                    cell_header.alignment = Alignment(horizontal="center")
            sheet.column_dimensions['C'].width = 7
            sheet.column_dimensions['D'].width = 31
            wb.save(file_path)
    else:
        if format == "Excel":
            excel_path = os.path.join("media","project","excel",f"{name[:-4]}.xlsx")
            file_path = excel_path
            AGS4.AGS4_to_excel(ags_file.ags_file.path,excel_path)
    if info:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.create_sheet(title="Basic information", index=0)
        # create fill color
        fill_cell = PatternFill(patternType='solid', fgColor='FFFF00') 
        # writing to the specified cell
        sheet.cell(row = 1, column = 1).value = 'This AGS file contains the following data:'
        sheet.cell(row = 1, column = 1).font = Font(bold = True )

        row = 2
        size = str(np.round(os.stat(ags_file.ags_file.path).st_size/1024,2)) + " kB"
        now = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
        try:
            version = sheets["TRAN"]["TRAN_AGS"][2]
        except:
            version = sheets["PROJ"]["PROJ_AGS"][1]
        basic_info = [("AGS File Name",name),("File Size",f"{size}"),("Time (UTC)",now),("AGS Version",f"{version}")
        ,("Project ID",Project_ID),("Project Name",Project_Name),("Project Location",Project_Location)
        ,("Client",Client),("Originator",Originator)]

        for i in range(len(basic_info)):
            sheet.cell(row = row, column = 1).value = basic_info[i][0]
            sheet.cell(row = row, column = 2).value = basic_info[i][1]
            row +=1

        sheet.cell(row = row, column = 2).value = "Value"
        sheet.cell(row = row, column = 2).font = Font(bold = True )
        sheet.cell(row = row, column = 3).value = "Units"
        sheet.cell(row = row, column = 3).font = Font(bold = True )

        for col in range(1, 10):
                cell_header = sheet.cell(row, col)
                cell_header.fill = fill_cell  
        row +=1
        basic_info_2 =[("No. of Borehole" ,N_Borehole ,"Number")
                    ,("Average Ground Level" ,Average_Ground_Level ,"m"),
        ("Average Water Level" , Average_Water_Level,"m"),("Max Borehole depth",Max_Borehole_depth,"m")]
        for i in range(len(basic_info_2)):
            sheet.cell(row = row, column = 1).value = basic_info_2[i][0]
            sheet.cell(row = row, column = 2).value = basic_info_2[i][1]
            sheet.cell(row = row, column = 3).value = basic_info_2[i][2]
            row +=1
        sheet.column_dimensions['A'].width = 37
        sheet.column_dimensions['B'].width = 40
        for row in range(11, 16):
                cell_header = sheet.cell(row, 2)
                cell_header.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['C'].width = 7
        sheet.column_dimensions['D'].width = 31
        wb.save(file_path)

    if record:
        df,elevation = get_main_df(sheets,id_name,coordinates_columns)
        user = settings.DATABASES['default']['USER']
        password = settings.DATABASES['default']['PASSWORD'].replace("@","%40")
        database_name = settings.DATABASES['default']['NAME']
        database_url = f'postgresql://{user}:{password}@localhost:5432/{database_name}'
        engine = create_engine(database_url,connect_args={'connect_timeout': 10}, echo=False)
        conn = engine.raw_connection()
        cur = conn.cursor()

        df_ispt = get_df('ISPT',["ISPT_NVAL","ISPT_PEN3",
                                "ISPT_PEN4","ISPT_PEN5","ISPT_PEN6"]
                         ,sheets,id_name,df,elevation,ags_file)
        df_ispt.head(0).to_sql('home_ispt', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_ispt.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_ispt', null="")
    
        df_CORE= get_df('CORE',["CORE_PREC","CORE_SREC","CORE_RQD","CORE_DIAM"]
                        ,sheets,id_name,df,elevation,ags_file)
        df_CORE.head(0).to_sql('home_core', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_CORE.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_core', null="")
    
        df_GCHM = get_df('GCHM',["GCHM_RESL","GCHM_NAME"]
                        ,sheets,id_name,df,elevation,ags_file)
        df_GCHM = df_GCHM.replace("Chloride contentof soil","Chloride_content_of_soil")
        df_GCHM = df_GCHM.replace("Chloride content of soil","Chloride_content_of_soil")
        df_GCHM = df_GCHM.replace("Sulphate content of soil","Sulphate_content_of_soil")
        df_GCHM = df_GCHM.replace("PH value","PH_Value")
        df_GCHM = df_GCHM.replace("Ph Value","PH_Value")
        df_GCHM = df_GCHM.replace("PH Value","PH_Value")
        df_GCHM = df_GCHM.replace("<0.002",0.002)
        df_GCHM = df_GCHM.pivot(index=["borhole_id","elevation","depth","machine"], columns='GCHM_NAME',
                                values='GCHM_RESL').reset_index().dropna().astype(float,errors="ignore")
        df_GCHM.head(0).to_sql('home_gchm', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_GCHM.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_gchm', null="")

        df_RUCS = get_df('RUCS',["RUCS_UCS","RUCS_E"]
                        ,sheets,id_name,df,elevation,ags_file)
        df_RUCS.head(0).to_sql('home_rucs', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_RUCS.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_rucs', null="")

        df_SHBG = get_df('SHBG',["SHBG_PCOH","SHBG_PHI","SHBG_RCOH","SHBG_RPHI"]
                        ,sheets,id_name,df,elevation,ags_file)
        df_SHBG.head(0).to_sql('home_shbg', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_SHBG.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_shbg', null="")

        df_GRAG = get_df('GRAG',["GRAG_GRAV"]
                        ,sheets,id_name,df,elevation,ags_file)
        df_GRAG.head(0).to_sql('home_grag', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_GRAG.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_grag', null="")
    
        df_GRAT = get_df('GRAT',["SPEC_DPTH","GRAT_SIZE","GRAT_PERP"]
                        ,sheets,id_name,df,elevation,ags_file)
        df_GRAT.head(0).to_sql('home_grat', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_GRAT.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_grat', null="")
    
        df_LLPL = get_df('LLPL',["LLPL_LL","LLPL_PL","LLPL_PI","LLPL_425"]
                        ,sheets,id_name,df,elevation,ags_file).replace(["NP","np"],0).astype(float,errors="ignore")
        df_LLPL.head(0).to_sql('home_llpl', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_LLPL.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_llpl', null="")

    
        df_LNMC = get_df('LNMC',["LNMC_MC"],sheets,id_name,df,elevation,ags_file)
        df_LNMC.head(0).to_sql('home_lnmc', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_LNMC.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_lnmc', null="")
    
        df_LPDN = get_df('LPDN',["LPDN_PDEN"],sheets,id_name,df,elevation,ags_file)
        df_LPDN.head(0).to_sql('home_lpdn', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_LPDN.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_lpdn', null="")
        
        df_RDEN = get_df('RDEN',["RDEN_BDEN"],sheets,id_name,df,elevation,ags_file)
        df_RDEN.head(0).to_sql('home_rden', engine, if_exists='replace',index=False)
        output = io.StringIO()
        df_RDEN.to_csv(output, sep='\t', header=False, index=False)
        output.seek(0)
        cur.copy_from(output, 'home_rden', null="")
        conn.commit()
        cur.close()
        conn.close()
    
    project_info = (Project_ID,Project_Name,Project_Location,Client,Originator)
    basic_info = (N_Borehole,Average_Ground_Level,Average_Water_Level,Max_Borehole_depth,total_drilled)
    return file_path , project_info ,basic_info
    


def check_ags (ags_file):
    
    path = os.path.join(os.path.dirname(ags_file.ags_file.path),"edited_ags")
    isExist = os.path.exists(path)
    if not isExist:
        os.makedirs(path)

    #editing the ags file
    id_name = 'LOCA_ID'
    with open(ags_file.ags_file.path,"r") as file:
        lines = file.readlines()
        if lines[0].find("**")+1:
            id_name ="HOLE_ID"
            for i in range(len(lines)):
                if len(lines[i]) > 1:
                    if lines[i].find("**")+1:
                        lines[i]='"GROUP",'+f'"{lines[i][3:-2]}"'+'\n'
                    if lines[i].find("*")+1:
                        lines[i]=lines[i].replace("*","")
                    if lines[i-1].split(",")[0] == '"GROUP"':
                        lines[i]='"HEADING",'+lines[i]
                    if lines[i-1].split(",")[0] == '"HEADING"' or lines[i-1].split(",")[0] == '"DATA"':
                        lines[i]='"DATA",'+lines[i]
        else:
            for i in range(len(lines)):
                if len(lines[i]) > 1:
                    if lines[i][-2] != '"':
                        lines[i]=lines[i].strip() + " "

    # storing the edited ones
    name = os.path.split(ags_file.ags_file.path)[1]
    ags_edited_path = os.path.join(path,name)
    with open(ags_edited_path,"w+") as edited_file:
        edited_file.writelines(lines)
        

    tables, _ = AGS4.AGS4_to_dataframe(ags_edited_path)
    
    groundWater_sheet = [i for i in tables.keys() if i in ["WSTG","MOND","WSTK"]][0]
    groundWater_dictionary = {"WSTG":"WSTG_DPTH","MOND":"MOND_RDNG","WSTK":"WSTK_DEP"}
    groundWater = groundWater_dictionary[groundWater_sheet]
    
    if id_name != 'LOCA_ID':
        common_sheet = "HOLE"
        Max_Borehole_depth_col = "HOLE_FDEP"
    else:
        common_sheet = "LOCA"
        Max_Borehole_depth_col = "LOCA_FDEP"
        
    df = pd.merge(tables[common_sheet],tables[groundWater_sheet], on=[id_name] ,how="outer")
    df = df.replace(["X","DATA"],np.nan)
    if id_name != 'LOCA_ID':
        df = df.drop(0)
    else:
        df = df.drop([0,1])
    df = df.replace([""],np.nan).dropna(axis=1, thresh=5)
    df = df.replace(["-"],0)
    for column in df.columns:
        df[column] = pd.to_numeric(df[column],errors='ignore')
    elevation_column = re.findall("\w+_GL|\w+_LOCZ"," ".join(df.columns))[0]
    
    # basic  information
    size = str(np.round(os.stat(ags_file.ags_file.path).st_size/1024,2)) + " kB"
    now = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    try:
        version = tables["TRAN"]["TRAN_AGS"][2]
    except:
        version = tables["PROJ"]["PROJ_AGS"][1]
        
    Project_ID = tables["PROJ"]["PROJ_ID"].iloc[-1]
    Project_Name = tables["PROJ"]["PROJ_NAME"].iloc[-1]
    Project_Location = tables["PROJ"]["PROJ_LOC"].iloc[-1]
    Client = tables["PROJ"]["PROJ_CLNT"].iloc[-1]
    try:
        Originator = tables["PROJ"]["PROJ_CONT"].iloc[-1]
    except:
        Originator = tables["PROJ"]["PROJ_ENG"].iloc[-1]
    


    Laboratory = ['GRAG','LLPL','LNMC','LPDN','RDEN','RELD','TREG','RPLT','RUCS','SHBG',"CLSS"]
    test_names = {'GRAG':"Gradation Tests",'LLPL':"Aterrberg Tests",'CLSS':"Aterrberg Tests",
     'LNMC':"Water/Moisture Content",'LPDN':"Specific Gravity",
     'RDEN':"Bulk Density",'RELD':"Relative Density",
     'TREG':"Triaxial Tests (Effective Stress)",'RPLT':"Point Load Testing",
     'RUCS':"UCS Tests",'SHBG':"Shear Box Testing"}
    test_sheets = [i for i in tables.keys() if i in Laboratory]
    Laboratory_Test_Data =[]
    
    if id_name == 'LOCA_ID':
        for i in test_sheets:
            Laboratory_Test_Data.append([test_names[i], tables[i]["SAMP_TOP"].count()-2,
                                        "Number",tables[i][f"{i}_METH"].iloc[-1]])
    else:
        for i in test_sheets:
            Laboratory_Test_Data.append([test_names[i], tables[i]["SAMP_TOP"].count()-1,"Number",""])
        if "ROCK" in tables.keys():
            Laboratory_Test_Data.append(["Water/Moisture Content", tables['ROCK']["ROCK_MC"].replace("",np.nan).count()-1,
                                        "Number",""])
            Laboratory_Test_Data.append(["UCS Tests", tables['ROCK']["ROCK_UCS"].replace("",np.nan).count()-1,
                                        "Number",""])
            Laboratory_Test_Data.append(["Bulk Density", tables['ROCK']["ROCK_BDEN"].replace("",np.nan).count()-1,
                                        "Number",""])
            

    chemical_tests = {"CL" : "Chloride Content","CO2": "Carbonate Content" ,"PH":"PH Value", "WS" : "Sulphate Content"
                     ,"PHS":"PH Value",'CACO3': "Carbonate Content", "SO4" : "Sulphate Content"}
    Chemical_Test_Data = []
    if 'GCHM' in tables.keys():
        for i in ["CL"  , "WS" , "CO2", "PH"]:
            Chemical_Test_Data.append([chemical_tests[i] , tables['GCHM'][tables['GCHM']["GCHM_CODE"]==i]["GCHM_CODE"].count()
                                      ,"Number",tables['GCHM']["GCHM_METH"].iloc[-1]])
    if 'CNMT' in tables.keys():
        for i in ['CACO3', 'CL', 'PHS', 'SO4']:
            Chemical_Test_Data.append([chemical_tests[i] , tables['CNMT'][tables['CNMT']["CNMT_TYPE"]==i]["CNMT_TYPE"].count()
                                      ,"Number",""])
    
    
    error=AGS4.check_file(ags_edited_path)
    wb = openpyxl.Workbook()
    sheet = wb.active
    row =1
    meta = error.pop("Metadata")
    sheet.cell(row = row, column = 1).value = "Metadata"
    row +=1
    for line in meta:
        sheet.cell(row = row, column = 1).value = line["line"]
        sheet.cell(row = row, column = 2).value = line['desc']
        row +=1
    total_errors = 0
    for rule in error.keys():
        sheet.cell(row = row, column = 1).value = rule
        row +=1
        sheet.cell(row = row, column = 1).value = "line"
        sheet.cell(row = row, column = 2).value = 'group'
        sheet.cell(row = row, column = 3).value = 'desc'
        row +=1
        for line in error[rule]:
            total_errors += 1
            sheet.cell(row = row, column = 1).value = line["line"]
            sheet.cell(row = row, column = 2).value = line['group']
            sheet.cell(row = row, column = 3).value = line['desc']
            row +=1
    
    sheet.cell(row=row+1, column=1).value =  "AGS file converter:"
    sheet.cell(row=row+1, column=2).value = '=HYPERLINK("{}", "{}")'.format(f"https://orycta.com/tools/AGSValidator", "AGS Validator")
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['C'].width = 140
    
    isExist = os.path.exists(os.path.join("media","project","excel"))
    if not isExist:
        os.makedirs(os.path.join("media","project","excel"))
    excel_path_error = os.path.join("media","project","excel",f"{name[:-4]}_error.xlsx")
    wb.save(excel_path_error)
    
    # Call a Workbook() function of openpyxl to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet  
    sheet = wb.active
    # create fill color
    fill_cell = PatternFill(patternType='solid', fgColor='FFFF00') 
    fill_cell_2 = PatternFill(patternType='solid', fgColor='F8CBAD') 
    # writing to the specified cell
    sheet.cell(row = 1, column = 1).value = 'This AGS file contains the following data:'
    sheet.cell(row = 1, column = 1).font = Font(bold = True )

    row = 2


    basic_info = [("AGS File Name",name),("File Size",f"{size}"),("Time (UTC)",now),("AGS Version",version)
     ,("Project ID",Project_ID),("Project Name",Project_Name),("Project Location",Project_Location)
     ,("Client",Client),("Originator",Originator)]

    for i in range(len(basic_info)):
        sheet.cell(row = row, column = 1).value = basic_info[i][0]
        sheet.cell(row = row, column = 2).value = basic_info[i][1]
        row +=1

    sheet.cell(row = row, column = 2).value = "Value"
    sheet.cell(row = row, column = 2).font = Font(bold = True )
    sheet.cell(row = row, column = 3).value = "Units"
    sheet.cell(row = row, column = 3).font = Font(bold = True )

    for col in range(1, 10):
            cell_header = sheet.cell(row, col)
            cell_header.fill = fill_cell  
    row +=1
    N_Borehole = df[id_name].count()
    Average_Ground_Level = np.round(np.average(df[elevation_column]),2)
    Average_Water_Level = np.round(np.average(df[groundWater]),2)
    Max_Borehole_depth = df[Max_Borehole_depth_col].max()
    basic_info_2 =[("No. of Borehole" ,N_Borehole ,"Number")
                   ,("Average Ground Level" ,Average_Ground_Level ,"m"),
    ("Average Water Level" , Average_Water_Level,"m"),("Max Borehole depth",Max_Borehole_depth,"m")]
    for i in range(len(basic_info_2)):
        sheet.cell(row = row, column = 1).value = basic_info_2[i][0]
        sheet.cell(row = row, column = 2).value = basic_info_2[i][1]
        sheet.cell(row = row, column = 3).value = basic_info_2[i][2]
        row +=1

    sheet.cell(row = row, column = 1).value = "Field Test Data"
    sheet.cell(row = row, column = 1).font = Font(bold = True )

    for col in range(1, 10):
            cell_header = sheet.cell(row, col)
            cell_header.fill = fill_cell  
    row +=1

    Field_Test_Data = [("CORE_PREC" , tables["CORE"]["CORE_PREC"].count()-1-int(id_name == 'LOCA_ID'),"Number")
    ,("CORE_SREC" , tables["CORE"]["CORE_SREC"].count()-1-int(id_name == 'LOCA_ID'),"Number")
    ,("CORE_RQD" , tables["CORE"]["CORE_RQD"].count()-1-int(id_name == 'LOCA_ID'),"Number")
    ,("N SPT Data" , tables["ISPT"]["ISPT_NVAL"].count()-1-int(id_name == 'LOCA_ID'),"Number")]
    for i in range(len(Field_Test_Data)):
        sheet.cell(row = row, column = 1).value = Field_Test_Data[i][0]
        sheet.cell(row = row, column = 2).value = Field_Test_Data[i][1]
        sheet.cell(row = row, column = 3).value = Field_Test_Data[i][2]
        row +=1


    sheet.cell(row = row, column = 1).value = "Laboratory Test Data"
    sheet.cell(row = row, column = 1).font = Font(bold = True )
    sheet.cell(row = row, column = 4).value = "Test reference"
    sheet.cell(row = row, column = 4).font = Font(bold = True )

    for col in range(1, 10):
            cell_header = sheet.cell(row, col)
            cell_header.fill = fill_cell  
    row +=1
    for i in range(len(Laboratory_Test_Data)):
        sheet.cell(row = row, column = 1).value = Laboratory_Test_Data[i][0]
        sheet.cell(row = row, column = 2).value = Laboratory_Test_Data[i][1]
        sheet.cell(row = row, column = 3).value = Laboratory_Test_Data[i][2]
        sheet.cell(row = row, column = 4).value = Laboratory_Test_Data[i][3]
        row +=1

    sheet.cell(row = row, column = 1).value = "Chemical Test Data"
    sheet.cell(row = row, column = 1).font = Font(bold = True )
    sheet.cell(row = row, column = 4).value = "Test reference"
    sheet.cell(row = row, column = 4).font = Font(bold = True )

    for col in range(1, 10):
            cell_header = sheet.cell(row, col)
            cell_header.fill = fill_cell  
    row +=1
    for i in range(len(Chemical_Test_Data)):
        sheet.cell(row = row, column = 1).value = Chemical_Test_Data[i][0]
        sheet.cell(row = row, column = 2).value = Chemical_Test_Data[i][1]
        sheet.cell(row = row, column = 3).value = Chemical_Test_Data[i][2]
        sheet.cell(row = row, column = 4).value = Chemical_Test_Data[i][3]
        row +=1

    sheet.cell(row = row, column = 1).value = "Error Report"
    sheet.cell(row = row, column = 1).font = Font(bold = True )


    for col in range(1, 10):
            cell_header = sheet.cell(row, col)
            cell_header.fill = fill_cell_2 
    row +=1
    # Add a hyperlink
    sheet.cell(row=row, column=1).value = '=HYPERLINK("{}", "{}")'.format(f"{name[:-4]}_error.xlsx", "Error file")
    
    sheet.cell(row=row+1, column=1).value =  "AGS file converter:"
    sheet.cell(row=row+1, column=2).value = '=HYPERLINK("{}", "{}")'.format(f"https://orycta.com/tools/AGSValidator", "AGS Validator")
    
    sheet.column_dimensions['A'].width = 37
    sheet.column_dimensions['B'].width = 40
    for row in range(11, 40):
            cell_header = sheet.cell(row, 2)
            cell_header.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['C'].width = 7
    sheet.column_dimensions['D'].width = 31

    

    # save the file
    excel_path_summary = os.path.join("media","project","excel",f"{name[:-4]}_summary.xlsx")
    wb.save(excel_path_summary)
    project_info = (name, size ,version ,Project_ID,Project_Name,Project_Location,Client,Originator)
    basic_info = (N_Borehole,Average_Ground_Level,Average_Water_Level,Max_Borehole_depth,total_errors)
    return excel_path_error,excel_path_summary ,project_info,basic_info














